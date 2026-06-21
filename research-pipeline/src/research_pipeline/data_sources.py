"""Track 2 — real point-in-time data loaders.

Three sources:
- **Ken French** (free) — daily factor / portfolio returns. Downloadable; may be cached
  under ``research-pipeline/data/`` (gitignored).
- **AQR data library** (free) — already-built long/short factor *return streams* per asset
  class (TSMOM by asset class; Value-and-Momentum-Everywhere factors). These are the inputs to
  the bounded cross-asset generalisation study (``crossasset.analyze_return_streams``).
- **WRDS / CRSP** (licensed) — the real equity cross-section. Read via the ``wrds`` package
  using YOUR credentials, or a local gitignored parquet extract.

HARD RULE (repo-wide): **licensed data is NEVER committed.** All cached/extracted data lives
under ``data/`` which is gitignored. These loaders never write to a tracked path.

VERIFICATION-SCOPE CAVEAT: the AQR loaders return *pre-built factor return streams*, not raw
price panels. They do NOT pass through the verified daily event-driven backtester, so results
derived from them are **breadth/generalisation evidence only** and must never be labelled
"verified no-look-ahead." That guarantee is scoped to the daily equity backtest. See
``crossasset.analyze_return_streams``.
"""

from __future__ import annotations

import datetime
import io
import pathlib
import urllib.request
import zipfile

import openpyxl
import pandas as pd

from .data import PricePanel

# research-pipeline/data — gitignored; never tracked.
DATA_DIR = pathlib.Path(__file__).resolve().parents[2] / "data"

# Ken French data library — each dataset is published as ``<name>_CSV.zip``.
_FRENCH_BASE = "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp"
_FRENCH_MISSING = (-99.99, -999.0, -99.99)  # sentinel codes for missing data


def _parse_french_csv(text: str) -> pd.DataFrame:
    """Parse the first data section of a Ken French CSV (decimal returns, NaN for missing).

    French files carry a text preamble, then one or more labelled sections (value-weighted,
    equal-weighted, ...). Each section is a comma header beginning with an empty first field
    followed by column names, then ``YYYYMMDD`` (daily) or ``YYYYMM`` (monthly) data rows. We
    take the first section, which is the one the dataset is named for (factors, or the
    value-weighted portfolio returns).
    """
    lines = text.splitlines()
    header_i = next(
        i for i, ln in enumerate(lines) if ln.startswith(",") and any(c.isalpha() for c in ln)
    )
    columns = [c.strip() for c in lines[header_i].split(",")[1:] if c.strip()]
    dates: list[str] = []
    rows: list[list[float]] = []
    for ln in lines[header_i + 1 :]:
        tok = ln.split(",")
        key = tok[0].strip()
        if not (key.isdigit() and len(key) in (6, 8)):
            break  # end of the first section
        dates.append(key)
        rows.append([float(x) for x in tok[1 : len(columns) + 1]])
    fmt = "%Y%m%d" if len(dates[0]) == 8 else "%Y%m"
    df = pd.DataFrame(rows, columns=columns, index=pd.to_datetime(dates, format=fmt))
    df = df.mask(df.isin(_FRENCH_MISSING))
    return df / 100.0  # French returns are in percent


def load_ken_french_factors(
    dataset: str = "F-F_Research_Data_Factors_daily",
    start: str | None = None,
    end: str | None = None,
) -> pd.DataFrame:
    """Ken French returns (decimal), fetched directly from Dartmouth (stdlib only, no deps).

    ``dataset`` is the library name, e.g. ``"49_Industry_Portfolios_daily"`` or
    ``"F-F_Research_Data_Factors_daily"``; the loader appends ``_CSV.zip``. Missing values
    (``-99.99``/``-999``) become ``NaN``. Optional ``start``/``end`` (``"YYYY-MM-DD"``) slice
    the date range. Used for real-data studies and the factor-attribution stage.
    """
    url = f"{_FRENCH_BASE}/{dataset}_CSV.zip"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:  # noqa: S310 - fixed https host
        payload = resp.read()
    archive = zipfile.ZipFile(io.BytesIO(payload))
    text = archive.read(archive.namelist()[0]).decode("latin-1")
    df = _parse_french_csv(text)
    if start is not None:
        df = df.loc[df.index >= pd.Timestamp(start)]
    if end is not None:
        df = df.loc[df.index <= pd.Timestamp(end)]
    return df


def load_crsp_daily(
    permnos_or_tickers: list[str] | None = None,
    start: str = "2005-01-01",
    end: str = "2023-12-31",
    wrds_username: str | None = None,
    local_parquet: str | pathlib.Path | None = None,
) -> PricePanel:
    """Daily CRSP adjusted-price panel as a ``PricePanel``.

    LICENSED — never commit the output. Two paths:
    1. ``local_parquet`` — a gitignored extract you already pulled (preferred for reuse).
    2. live ``wrds`` connection (``pip install wrds``; needs your WRDS login).

    Builds total-return-adjusted prices (``prc / cfacpr``) pivoted to dates x permno.
    """
    if local_parquet is not None:
        df = pd.read_parquet(local_parquet)
    else:
        try:
            import wrds  # type: ignore[import-not-found]
        except ImportError as exc:  # pragma: no cover - environment-dependent
            raise ImportError(
                "pip install wrds and provide credentials, or pass local_parquet"
            ) from exc
        db = wrds.Connection(wrds_username=wrds_username)
        query = (
            "select date, permno, abs(prc)/cfacpr as adj_price "
            "from crsp.dsf "
            f"where date between '{start}' and '{end}'"
        )
        df = db.raw_sql(query, date_cols=["date"])
        db.close()
    prices = df.pivot(index="date", columns="permno", values="adj_price").sort_index()
    prices.columns = [str(c) for c in prices.columns]
    return PricePanel(prices)


# --------------------------------------------------------------------------------------------
# AQR data library (free) — pre-built long/short factor RETURN streams per asset class.
# https://www.aqr.com/Insights/Datasets . Each dataset is a single .xlsx with a text preamble,
# a header row, then month-end date rows and decimal-return factor columns.
# --------------------------------------------------------------------------------------------

_AQR_MEDIA = "https://www.aqr.com/-/media/AQR/Documents/Insights/Data-Sets"
AQR_TSMOM_URL = f"{_AQR_MEDIA}/Time-Series-Momentum-Factors-Monthly.xlsx"
AQR_VME_MONTHLY_URL = f"{_AQR_MEDIA}/Value-and-Momentum-Everywhere-Factors-Monthly.xlsx"


def _parse_aqr_date(value: object) -> pd.Timestamp | None:
    """Parse an AQR date cell (a ``datetime`` or an ``MM/DD/YYYY`` string) to a Timestamp."""
    if isinstance(value, datetime.datetime):
        return pd.Timestamp(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return pd.Timestamp(datetime.datetime.strptime(s, "%m/%d/%Y"))
        except ValueError:
            try:
                return pd.Timestamp(s)
            except (ValueError, TypeError):
                return None
    return None


def parse_aqr_sheet(rows: list[tuple[object, ...]]) -> pd.DataFrame:
    """Parse the rows of an AQR factor sheet into a tidy decimal-return DataFrame.

    AQR factor files have a multi-line text preamble, a header row of factor names, then
    month-end rows of *already-decimal* returns. The header is located as the last text row
    immediately above the first row whose first cell parses as a date; its non-empty labels
    (after the leading date column) become the columns. Blank cells become ``NaN`` and trailing
    all-blank rows end the data block.

    Pure function over an in-memory row list (the network fetch lives in the loaders), so it is
    unit-testable against a small synthetic fixture without hitting AQR.
    """
    first_data: int | None = None
    for i, row in enumerate(rows):
        if row and _parse_aqr_date(row[0]) is not None:
            first_data = i
            break
    if first_data is None or first_data == 0:
        raise ValueError("could not locate the AQR data block (no date row found)")
    header_i = first_data - 1
    header = rows[header_i]
    # Columns are the non-empty labels after the leading date column. Keep their positions so a
    # gap in the header (AQR pads with blank columns) does not shift the data alignment.
    col_positions = [j for j in range(1, len(header)) if str(header[j] or "").strip()]
    columns = [str(header[j]).strip() for j in col_positions]
    if not columns:
        raise ValueError("AQR header row carried no factor columns")
    dates: list[pd.Timestamp] = []
    data: list[list[float]] = []
    for row in rows[first_data:]:
        ts = _parse_aqr_date(row[0]) if row else None
        if ts is None:
            break  # trailing blank / footnote rows end the data block
        out: list[float] = []
        for j in col_positions:
            cell = row[j] if j < len(row) else None
            out.append(float(cell) if isinstance(cell, (int, float)) else float("nan"))
        dates.append(ts)
        data.append(out)
    df = pd.DataFrame(data, columns=columns, index=pd.DatetimeIndex(dates))
    return df.sort_index()


def _fetch_aqr_sheet(url: str, sheet: str) -> pd.DataFrame:
    """Download an AQR .xlsx and parse the named factor sheet to decimal returns."""
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:  # noqa: S310 - fixed https host
        payload = resp.read()
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet {sheet!r} not in workbook (sheets: {wb.sheetnames})")
    rows = list(wb[sheet].iter_rows(values_only=True))
    return parse_aqr_sheet(rows)


# AQR's per-asset-class TSMOM column suffixes -> readable asset-class names.
_TSMOM_ASSET_CLASSES = {
    "TSMOM^EQ": "equities",
    "TSMOM^FI": "fixed_income",
    "TSMOM^FX": "currencies",
    "TSMOM^CM": "commodities",
}


def load_aqr_tsmom(asset_classes_only: bool = True) -> pd.DataFrame:
    """AQR Time-Series Momentum factors, monthly (free), as decimal return streams.

    Columns are the long/short TSMOM factor returns: a diversified ``TSMOM`` plus one stream
    per asset class (``TSMOM^EQ`` equities, ``TSMOM^FI`` fixed income, ``TSMOM^FX`` currencies,
    ``TSMOM^CM`` commodities). With ``asset_classes_only=True`` (default) the four asset-class
    streams are returned with readable names, which is the input the cross-asset study wants.

    Returns are AQR's published *excess returns* in decimal form; the index is month-end dates.
    NOT routed through the verified backtester — generalisation evidence only.
    """
    df = _fetch_aqr_sheet(AQR_TSMOM_URL, "TSMOM Factors")
    if asset_classes_only:
        present = [c for c in _TSMOM_ASSET_CLASSES if c in df.columns]
        df = df[present].rename(columns=_TSMOM_ASSET_CLASSES)
    return df.dropna(how="all")


# AQR's "all asset classes" VME long/short factor columns -> readable names. Momentum is the
# signal the study holds fixed across classes; value is loaded too for the correlation table.
_VME_MOMENTUM_CLASSES = {
    "MOMLS_VME_US90": "equities_us",
    "MOMLS_VME_UK90": "equities_uk",
    "MOMLS_VME_ROE90": "equities_europe",
    "MOMLS_VME_JP90": "equities_japan",
    "MOMLS_VME_EQ": "equity_indices",
    "MOMLS_VME_FX": "currencies",
    "MOMLS_VME_FI": "fixed_income",
    "MOMLS_VME_COM": "commodities",
}


def load_aqr_vme_monthly(momentum_only: bool = True) -> pd.DataFrame:
    """AQR Value-and-Momentum-Everywhere factors, monthly (free), as decimal return streams.

    Columns are zero-cost long/short factor returns across eight markets/asset classes. With
    ``momentum_only=True`` (default) the eight per-asset-class *momentum* streams are returned
    with readable names so the same momentum effect can be compared across classes; set it
    ``False`` to get every raw column (value + momentum, global averages, regional equities).

    Returns are decimal; the index is month-end dates. NOT routed through the verified
    backtester — generalisation evidence only.
    """
    df = _fetch_aqr_sheet(AQR_VME_MONTHLY_URL, "VME Factors")
    if momentum_only:
        present = [c for c in _VME_MOMENTUM_CLASSES if c in df.columns]
        df = df[present].rename(columns=_VME_MOMENTUM_CLASSES)
    return df.dropna(how="all")
